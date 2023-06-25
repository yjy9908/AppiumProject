import os.path
import pandas as pd
from openpyxl import load_workbook,workbook


def read_data_from_excel(excel_file,sheet_name):
    return_value=[]

    #判断文件是否存在
    if not os.path.exists(excel_file):
        raise ValueError("File not exists")

    #打开指定的sheet
    wb=load_workbook(excel_file)

    #按照pytest接受的格式输出数据
    for s in wb.sheetnames:
        if s== sheet_name:
            sheet=wb[sheet_name]
            for row in sheet.rows:
                return_value.append([col.value for col in row])

    return return_value[1:]

